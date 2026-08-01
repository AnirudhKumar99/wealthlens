"""
excel_importer.py — Flexible, Robust Excel Parser & Importer for WealthLens 2.0
Supports importing both WealthLens exported reports and user-created custom Excel files.
"""
import uuid
import io
import re
from typing import Optional, Any
import openpyxl

from database import (
    get_conn, create_profile, insert_item,
    get_profiles_by_user_id
)


def _cell_val(c: Any) -> Any:
    """Safely extract cell value whether object is a Cell, ReadOnlyCell, or primitive."""
    if c is None:
        return None
    if hasattr(c, 'value'):
        return c.value
    return c


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Safely convert a cell value to float."""
    v = _cell_val(val)
    if v is None:
        return default
    try:
        if isinstance(v, str):
            v = v.strip().replace('%', '').replace(',', '').replace('₹', '').replace('$', '')
            if not v or v in ('-', 'N/A', 'none', 'null'):
                return default
        return float(v)
    except (ValueError, TypeError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    """Safely convert a cell value to int."""
    v = _cell_val(val)
    if v is None:
        return default
    try:
        if isinstance(v, str):
            v = v.strip().replace(',', '').replace('₹', '').replace('$', '')
            if not v or v in ('-', 'N/A', 'none', 'null'):
                return default
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _safe_str(val: Any, default: str = '') -> str:
    """Safely convert a cell value to string."""
    v = _cell_val(val)
    if v is None:
        return default
    return str(v).strip()


def _clean_header(header_name: str) -> str:
    """Clean emojis, currencies, symbols from header names."""
    h = _safe_str(header_name).lower().strip()
    for ch in ['🟣', '🔵', '🔷', '🟡', '💼', '📈', '🎯', '🏦', '🛡️', '👤', '📊', '📅', '₹', '$', '%', '(', ')']:
        h = h.replace(ch, '')
    return re.sub(r'\s+', ' ', h).strip()


def _find_header_row(ws, max_scan_rows=20):
    """Find the row index where headers start (first row with 2+ non-empty cells)."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        non_empty = sum(1 for cell in row if _safe_str(cell))
        if non_empty >= 2:
            return row_idx
    return None


def _get_flexible_value(row_dict: dict, keys: list, default=None):
    """Look up a value in row_dict matching any key in keys (fuzzy match)."""
    for k in keys:
        k_clean = _clean_header(k)
        for r_key, r_val in row_dict.items():
            r_clean = _clean_header(r_key)
            if k_clean == r_clean or k_clean in r_clean or r_clean in k_clean:
                if r_val is not None and str(r_val).strip() != '':
                    return r_val
    return default


def _read_table(ws, start_row=None):
    """Read a table from a worksheet starting at the header row.
    Returns a list of dicts keyed by cleaned header names."""
    if start_row is None:
        start_row = _find_header_row(ws)
    if start_row is None:
        return []

    header_values = []
    for r in ws.iter_rows(min_row=start_row, max_row=start_row, values_only=True):
        for cell in r:
            header_values.append(_clean_header(cell))

    if not any(header_values):
        return []

    rows = []
    for row in ws.iter_rows(min_row=start_row + 1, values_only=True):
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        row_dict = {}
        for idx, val in enumerate(row):
            if idx < len(header_values) and header_values[idx]:
                row_dict[header_values[idx]] = val
        if row_dict and any(v is not None and str(v).strip() != '' for v in row_dict.values()):
            rows.append(row_dict)
    return rows


def _find_section_start(ws, section_keywords: list):
    """Find the row where a table section starts by looking for keywords in the header row."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        clean_row = [_clean_header(c) for c in row if c is not None]
        if not clean_row:
            continue
        for kw in section_keywords:
            kw_clean = _clean_header(kw)
            if any(kw_clean in cell_str for cell_str in clean_row):
                # Ensure this is a header row with multiple non-empty cells
                if sum(1 for c in clean_row if c) >= 2:
                    return row_idx
    return None


def parse_family_excel(file_bytes: bytes) -> dict:
    """
    Parse a Family Household Master Report or custom Excel file.
    Returns structured data ready for database insertion.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {
        "members": [],
        "assets": [],
        "sips": [],
        "goals": [],
        "loans": [],
        "insurance": []
    }

    # 1. Parse Members / Executive Summary
    summary_sheet_name = None
    for name in wb.sheetnames:
        if any(k in name.lower() for k in ['executive', 'summary', 'member', 'profile', 'overview']):
            summary_sheet_name = name
            break

    if summary_sheet_name:
        ws = wb[summary_sheet_name]
        member_start = _find_section_start(ws, ['relationship', 'retirement target', 'role'])
        if member_start is None:
            member_start = _find_header_row(ws)

        if member_start:
            members_data = _read_table(ws, start_row=member_start)
            for m in members_data:
                name_val = _safe_str(_get_flexible_value(m, ['family member', 'member name', 'name', 'member']))
                if not name_val or any(k in name_val.lower() for k in ['total', 'summary', 'household', 'kpi', 'family member', 'metric']):
                    continue
                result["members"].append({
                    "family_name": name_val,
                    "role": _safe_str(_get_flexible_value(m, ['relationship / role', 'relationship', 'role'], 'Family Member')),
                    "current_age": _safe_int(_get_flexible_value(m, ['age', 'current age'], 35)),
                    "retirement_age": _safe_int(_get_flexible_value(m, ['retirement target', 'retirement age', 'retire at'], 60)),
                })

    # 2. Parse Assets
    assets_sheet_name = None
    for name in wb.sheetnames:
        if 'asset' in name.lower() or 'investment' in name.lower() or 'portfolio' in name.lower():
            assets_sheet_name = name
            break

    if assets_sheet_name:
        ws = wb[assets_sheet_name]
        assets_data = _read_table(ws)
        for a in assets_data:
            name_val = _safe_str(_get_flexible_value(a, ['asset name', 'name', 'asset', 'investment name', 'fund name', 'description']))
            if not name_val or any(k in name_val.lower() for k in ['total', 'summary', 'asset name', 'name']):
                continue
            owner = _safe_str(_get_flexible_value(a, ['family member', 'owner', 'member name', 'profile', 'member']))
            ret_rate = _safe_float(_get_flexible_value(a, ['expected return rate', 'return rate', 'return', 'cagr', 'roi', 'rate'], 8.0))
            if 0 < ret_rate < 1:
                ret_rate = ret_rate * 100

            result["assets"].append({
                "owner": owner,
                "name": name_val,
                "asset_class": _safe_str(_get_flexible_value(a, ['asset class', 'type', 'category', 'class'], 'equity')).lower(),
                "value": _safe_float(_get_flexible_value(a, ['current value', 'value', 'amount', 'balance', 'valuation'], 0.0)),
                "return_rate": ret_rate
            })

    # 3. Parse SIPs
    sips_sheet_name = None
    for name in wb.sheetnames:
        if 'sip' in name.lower() or 'recurring' in name.lower():
            sips_sheet_name = name
            break

    if sips_sheet_name:
        ws = wb[sips_sheet_name]
        sips_data = _read_table(ws)
        for s in sips_data:
            name_val = _safe_str(_get_flexible_value(s, ['sip / investment name', 'sip name', 'name', 'scheme name', 'fund name']))
            if not name_val or any(k in name_val.lower() for k in ['total', 'summary', 'sip name', 'name']):
                continue
            owner = _safe_str(_get_flexible_value(s, ['family member', 'owner', 'member name', 'profile', 'member']))
            step_up = _safe_float(_get_flexible_value(s, ['step-up rate', 'step_up_pct', 'step up', 'stepup'], 0.0))
            if 0 < step_up < 1:
                step_up = step_up * 100
            ret_rate = _safe_float(_get_flexible_value(s, ['return rate', 'return_rate', 'return', 'cagr', 'roi'], 12.0))
            if 0 < ret_rate < 1:
                ret_rate = ret_rate * 100

            end_y = _get_flexible_value(s, ['end year', 'end_year', 'end'])
            if end_y and str(end_y).lower().strip() in ('forever', 'none', '-', '', 'n/a'):
                end_y = None
            elif end_y:
                end_y = _safe_int(end_y)

            result["sips"].append({
                "owner": owner,
                "name": name_val,
                "asset_class": _safe_str(_get_flexible_value(s, ['asset class', 'type', 'category'], 'equity')).lower(),
                "monthly_amount": _safe_float(_get_flexible_value(s, ['monthly amount', 'monthly_amount', 'amount', 'monthly sip', 'sip amount'], 0.0)),
                "step_up_pct": step_up,
                "return_rate": ret_rate,
                "start_year": _safe_int(_get_flexible_value(s, ['start year', 'start_year', 'start'], 2026)),
                "end_year": end_y
            })

    # 4. Parse Goals
    goals_sheet_name = None
    for name in wb.sheetnames:
        if 'goal' in name.lower() or 'target' in name.lower():
            goals_sheet_name = name
            break

    if goals_sheet_name:
        ws = wb[goals_sheet_name]
        goals_data = _read_table(ws)
        for g in goals_data:
            name_val = _safe_str(_get_flexible_value(g, ['goal name', 'name', 'goal', 'target']))
            if not name_val or any(k in name_val.lower() for k in ['total', 'summary', 'goal name', 'name']):
                continue
            owner = _safe_str(_get_flexible_value(g, ['family member', 'owner', 'member name', 'profile', 'member']))
            inf_rate = _safe_float(_get_flexible_value(g, ['inflation rate', 'inflation_rate', 'inflation'], 6.0))
            if 0 < inf_rate < 1:
                inf_rate = inf_rate * 100

            result["goals"].append({
                "owner": owner,
                "name": name_val,
                "priority": _safe_str(_get_flexible_value(g, ['priority', 'importance'], 'need')).lower(),
                "present_value": _safe_float(_get_flexible_value(g, ['present value', 'present_value', 'value', 'amount', 'cost', 'target amount'], 0.0)),
                "target_year": _safe_int(_get_flexible_value(g, ['target year', 'target_year', 'year', 'due year'], 2030)),
                "goal_type": _safe_str(_get_flexible_value(g, ['goal type', 'goal_type', 'type'], 'lump_sum')).lower().replace(' ', '_'),
                "inflation_rate": inf_rate
            })

    # 5. Parse Insurance & Loans
    il_sheet_name = None
    for name in wb.sheetnames:
        if any(k in name.lower() for k in ['insurance', 'loan', 'liability', 'debt', 'policy']):
            il_sheet_name = name
            break

    if il_sheet_name:
        ws = wb[il_sheet_name]
        
        loans_start = _find_section_start(ws, ['loan name', 'principal', 'emis paid', 'loan type'])
        if loans_start is None:
            loans_start = _find_header_row(ws)

        if loans_start:
            loans_data = _read_table(ws, start_row=loans_start)
            for l in loans_data:
                name_val = _safe_str(_get_flexible_value(l, ['loan name', 'name', 'liability', 'bank / loan name']))
                if not name_val or any(k in name_val.lower() for k in ['total', 'summary', 'loan name', 'name', 'policy name', 'insurance']):
                    continue
                owner = _safe_str(_get_flexible_value(l, ['family member', 'owner', 'member name', 'profile', 'member']))
                roi = _safe_float(_get_flexible_value(l, ['roi', 'roi_pct', 'interest rate', 'rate'], 8.5))
                if 0 < roi < 1:
                    roi = roi * 100
                result["loans"].append({
                    "owner": owner,
                    "name": name_val,
                    "loan_type": _safe_str(_get_flexible_value(l, ['loan type', 'loan_type', 'type'], 'home')).lower(),
                    "principal": _safe_float(_get_flexible_value(l, ['principal', 'outstanding principal', 'amount', 'balance'], 0.0)),
                    "roi_pct": roi,
                    "total_months": _safe_int(_get_flexible_value(l, ['total months', 'total_months', 'tenure', 'months'], 240)),
                    "emis_paid": _safe_int(_get_flexible_value(l, ['emis paid', 'emis_paid', 'months paid'], 0))
                })

        ins_start = _find_section_start(ws, ['policy name', 'annual premium', 'death benefit', 'policy type'])
        if ins_start is None:
            ins_start = _find_header_row(ws)

        if ins_start:
            ins_data = _read_table(ws, start_row=ins_start)
            for ins in ins_data:
                name_val = _safe_str(_get_flexible_value(ins, ['policy name', 'name', 'insurance name', 'plan name']))
                if not name_val or any(k in name_val.lower() for k in ['total', 'summary', 'policy name', 'name', 'loan name']):
                    continue
                owner = _safe_str(_get_flexible_value(ins, ['family member', 'owner', 'member name', 'profile', 'member']))
                income_period = _safe_str(_get_flexible_value(ins, ['income period', 'period']))
                inc_start = None
                inc_end = None
                if ' - ' in income_period:
                    parts = income_period.split(' - ')
                    inc_start = _safe_int(parts[0]) if parts[0].strip() != '-' else None
                    inc_end = _safe_int(parts[1]) if len(parts) > 1 and parts[1].strip() != '-' else None

                result["insurance"].append({
                    "owner": owner,
                    "name": name_val,
                    "annual_premium": _safe_float(_get_flexible_value(ins, ['annual premium', 'annual_premium', 'premium'], 0.0)),
                    "premium_end_year": _safe_int(_get_flexible_value(ins, ['premium end year', 'premium_end_year', 'premium end'], 2035)),
                    "income_start_year": inc_start or _safe_int(_get_flexible_value(ins, ['income start year', 'income_start_year']), None),
                    "annual_income": _safe_float(_get_flexible_value(ins, ['annual income', 'annual_income', 'income'], 0.0)),
                    "income_end_year": inc_end or _safe_int(_get_flexible_value(ins, ['income end year', 'income_end_year']), None),
                    "death_benefit": _safe_float(_get_flexible_value(ins, ['death benefit', 'death_benefit', 'sum assured', 'cover'], 0.0))
                })

    wb.close()
    return result


def import_family_excel(user_id: str, file_bytes: bytes) -> dict:
    """
    Import a Family Household Master Excel report into the database.
    Creates/matches profiles and inserts all items.
    Returns summary stats.
    """
    parsed = parse_family_excel(file_bytes)

    existing_profiles = get_profiles_by_user_id(user_id)
    profile_map = {}
    for p in existing_profiles:
        profile_map[p["family_name"].strip().lower()] = p["id"]

    for member in parsed["members"]:
        key = member["family_name"].strip().lower()
        if key not in profile_map:
            new_id = str(uuid.uuid4())
            create_profile(new_id, {
                "user_id": user_id,
                "family_name": member["family_name"],
                "role": member.get("role", "Family Member"),
                "current_age": member.get("current_age", 35),
                "retirement_age": member.get("retirement_age", 60),
            }, seed_defaults=False)
            profile_map[key] = new_id

    all_owners = set()
    for section in [parsed["assets"], parsed["sips"], parsed["goals"], parsed["loans"], parsed["insurance"]]:
        for item in section:
            owner = item.get("owner", "").strip()
            if owner:
                all_owners.add(owner)

    for owner in all_owners:
        key = owner.strip().lower()
        if key not in profile_map:
            new_id = str(uuid.uuid4())
            create_profile(new_id, {
                "user_id": user_id,
                "family_name": owner,
                "role": "Family Member",
            }, seed_defaults=False)
            profile_map[key] = new_id

    if not profile_map:
        new_id = str(uuid.uuid4())
        create_profile(new_id, {
            "user_id": user_id,
            "family_name": "My Profile",
            "role": "Self",
        }, seed_defaults=False)
        profile_map["my profile"] = new_id

    def resolve_profile(owner_name):
        key = owner_name.strip().lower()
        if key in profile_map:
            return profile_map[key]
        for pname, pid in profile_map.items():
            if key in pname or pname in key:
                return pid
        return list(profile_map.values())[0]

    counts = {"assets": 0, "sips": 0, "goals": 0, "loans": 0, "insurance": 0, "profiles": len(profile_map)}

    for a in parsed["assets"]:
        pid = resolve_profile(a["owner"])
        insert_item("assets", {
            "id": str(uuid.uuid4()),
            "profile_id": pid,
            "name": a["name"],
            "asset_class": a["asset_class"],
            "value": a["value"],
            "return_rate": a["return_rate"]
        })
        counts["assets"] += 1

    for s in parsed["sips"]:
        pid = resolve_profile(s["owner"])
        data = {
            "id": str(uuid.uuid4()),
            "profile_id": pid,
            "name": s["name"],
            "asset_class": s["asset_class"],
            "monthly_amount": s["monthly_amount"],
            "step_up_pct": s["step_up_pct"],
            "return_rate": s["return_rate"],
            "start_year": s["start_year"],
        }
        if s.get("end_year"):
            data["end_year"] = s["end_year"]
        insert_item("sips", data)
        counts["sips"] += 1

    for g in parsed["goals"]:
        pid = resolve_profile(g["owner"])
        insert_item("goals", {
            "id": str(uuid.uuid4()),
            "profile_id": pid,
            "name": g["name"],
            "priority": g["priority"],
            "present_value": g["present_value"],
            "target_year": g["target_year"],
            "goal_type": g["goal_type"],
            "inflation_rate": g["inflation_rate"]
        })
        counts["goals"] += 1

    for l in parsed["loans"]:
        pid = resolve_profile(l["owner"])
        insert_item("loans", {
            "id": str(uuid.uuid4()),
            "profile_id": pid,
            "name": l["name"],
            "loan_type": l["loan_type"],
            "principal": l["principal"],
            "roi_pct": l["roi_pct"],
            "total_months": l["total_months"],
            "emis_paid": l["emis_paid"]
        })
        counts["loans"] += 1

    for ins in parsed["insurance"]:
        pid = resolve_profile(ins["owner"])
        data = {
            "id": str(uuid.uuid4()),
            "profile_id": pid,
            "name": ins["name"],
            "annual_premium": ins["annual_premium"],
            "premium_end_year": ins["premium_end_year"],
            "annual_income": ins["annual_income"],
            "death_benefit": ins["death_benefit"]
        }
        if ins.get("income_start_year"):
            data["income_start_year"] = ins["income_start_year"]
        if ins.get("income_end_year"):
            data["income_end_year"] = ins["income_end_year"]
        insert_item("insurance_plans", data)
        counts["insurance"] += 1

    return counts


def import_single_profile_excel(profile_id: str, file_bytes: bytes) -> dict:
    """
    Import an Excel report into a single existing profile.
    Reads all sheets and inserts items into the given profile.
    Returns summary stats.
    """
    parsed = parse_family_excel(file_bytes)
    counts = {"assets": 0, "sips": 0, "goals": 0, "loans": 0, "insurance": 0}

    for a in parsed["assets"]:
        insert_item("assets", {
            "id": str(uuid.uuid4()),
            "profile_id": profile_id,
            "name": a["name"],
            "asset_class": a["asset_class"],
            "value": a["value"],
            "return_rate": a["return_rate"]
        })
        counts["assets"] += 1

    for s in parsed["sips"]:
        data = {
            "id": str(uuid.uuid4()),
            "profile_id": profile_id,
            "name": s["name"],
            "asset_class": s["asset_class"],
            "monthly_amount": s["monthly_amount"],
            "step_up_pct": s["step_up_pct"],
            "return_rate": s["return_rate"],
            "start_year": s["start_year"],
        }
        if s.get("end_year"):
            data["end_year"] = s["end_year"]
        insert_item("sips", data)
        counts["sips"] += 1

    for g in parsed["goals"]:
        insert_item("goals", {
            "id": str(uuid.uuid4()),
            "profile_id": profile_id,
            "name": g["name"],
            "priority": g["priority"],
            "present_value": g["present_value"],
            "target_year": g["target_year"],
            "goal_type": g["goal_type"],
            "inflation_rate": g["inflation_rate"]
        })
        counts["goals"] += 1

    for l in parsed["loans"]:
        insert_item("loans", {
            "id": str(uuid.uuid4()),
            "profile_id": profile_id,
            "name": l["name"],
            "loan_type": l["loan_type"],
            "principal": l["principal"],
            "roi_pct": l["roi_pct"],
            "total_months": l["total_months"],
            "emis_paid": l["emis_paid"]
        })
        counts["loans"] += 1

    for ins in parsed["insurance"]:
        data = {
            "id": str(uuid.uuid4()),
            "profile_id": profile_id,
            "name": ins["name"],
            "annual_premium": ins["annual_premium"],
            "premium_end_year": ins["premium_end_year"],
            "annual_income": ins["annual_income"],
            "death_benefit": ins["death_benefit"]
        }
        if ins.get("income_start_year"):
            data["income_start_year"] = ins["income_start_year"]
        if ins.get("income_end_year"):
            data["income_end_year"] = ins["income_end_year"]
        insert_item("insurance_plans", data)
        counts["insurance"] += 1

    return counts
