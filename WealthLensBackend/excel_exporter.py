"""
excel_exporter.py — Multi-sheet Excel workbook generator for WealthLens 2.0
"""
import io
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_financial_excel_report(
    profile: dict,
    assets: list[dict],
    goals: list[dict],
    sips: list[dict],
    insurance_plans: list[dict],
    loans: list[dict],
    sim_result: dict
) -> io.BytesIO:
    """
    Generates a beautifully styled 6-sheet Excel workbook:
    1. Executive Summary
    2. Yearly Trajectory Schedule
    3. Assets Inventory
    4. SIPs & Investments
    5. Milestone Goals
    6. Loans & Insurance
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    title_font  = Font(name=font_family, size=16, bold=True, color="2D1B69")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="6B5B95")
    section_font = Font(name=font_family, size=12, bold=True, color="2D1B69")
    bold_font   = Font(name=font_family, size=11, bold=True, color="2D1B69")
    regular_font = Font(name=font_family, size=10, color="2D1B69")
    
    thin_border = Border(
        left=Side(style='thin', color='E0D7FF'),
        right=Side(style='thin', color='E0D7FF'),
        top=Side(style='thin', color='E0D7FF'),
        bottom=Side(style='thin', color='E0D7FF')
    )

    currency_fmt = '#,##0'
    pct_fmt = '0.0%'

    summary_data = sim_result.get("summary", {})
    yearly_data = sim_result.get("yearly_data", [])

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 1: Executive Summary
    # ─────────────────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet(title="Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    ws1["A1"] = "🔮 WealthLens — Financial Master Report"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Family Profile: {profile.get('family_name', 'My Family')} | Generated Financial Analysis"
    ws1["A2"].font = subtitle_font

    ws1["A4"] = "📌 Baseline Profile Parameters"
    ws1["A4"].font = section_font

    profile_headers = ["Parameter", "Value"]
    ws1.append([])
    ws1.append(profile_headers)
    p_header_row = ws1.max_row
    for col_num, h in enumerate(profile_headers, 1):
        cell = ws1.cell(row=p_header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    prof_rows = [
        ("Current Age", profile.get("current_age", 35)),
        ("Retirement Age", profile.get("retirement_age", 60)),
        ("Life Expectancy", profile.get("life_expectancy", 85)),
        ("Annual Income", profile.get("annual_income", 0)),
        ("Monthly Expenses in Retirement (PV)", profile.get("monthly_expenses_retirement", 40000)),
        ("Post-Retirement Inflation Rate (%)", f"{profile.get('retirement_inflation_rate', 7.0)}%"),
        ("Currency", profile.get("currency", "INR")),
    ]
    for param, val in prof_rows:
        ws1.append([param, val])
        r = ws1.max_row
        ws1.cell(row=r, column=1).font = regular_font
        ws1.cell(row=r, column=2).font = bold_font
        if isinstance(val, (int, float)):
            ws1.cell(row=r, column=2).number_format = currency_fmt

    ws1.append([])
    ws1.append(["📊 Key Simulation Analytics"])
    ws1.cell(row=ws1.max_row, column=1).font = section_font

    kpi_headers = ["Metric", "Value", "Status / Notes"]
    ws1.append(kpi_headers)
    k_header_row = ws1.max_row
    for col_num, h in enumerate(kpi_headers, 1):
        cell = ws1.cell(row=k_header_row, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    health_score = summary_data.get("health_score", 0)
    health_status = "🟢 Healthy" if health_score >= 70 else "🟡 Moderate" if health_score >= 40 else "🔴 At Risk"

    kpi_rows = [
        ("Wealth Health Score", f"{health_score} / 100", health_status),
        ("Total Starting Assets", sim_result.get("total_portfolio", 0), "Aggregated liquid & illiquid assets"),
        ("Blended Portfolio Return", f"{sim_result.get('blended_return', 0)}%", "Weighted annual yield across assets"),
        ("Portfolio at Retirement", summary_data.get("portfolio_at_retirement", 0), "Projected nest egg at retirement age"),
        ("FI Corpus Required (4% SWR)", summary_data.get("fi_corpus_needed", 0), "25x annual retirement expenses"),
        ("FI Ratio", f"{summary_data.get('fi_ratio', 0)}x", "Portfolio at retirement / FI target"),
        ("Portfolio Exhausted?", "YES" if summary_data.get("portfolio_exhausted") else "NO", f"Solvent through age {summary_data.get('life_expectancy', 85)}"),
    ]
    for m, v, s in kpi_rows:
        ws1.append([m, v, s])
        r = ws1.max_row
        ws1.cell(row=r, column=1).font = regular_font
        ws1.cell(row=r, column=2).font = bold_font
        ws1.cell(row=r, column=3).font = subtitle_font
        if isinstance(v, (int, float)):
            ws1.cell(row=r, column=2).number_format = currency_fmt

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 2: Yearly Trajectory Schedule
    # ─────────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Yearly Trajectory Schedule")
    ws2.views.sheetView[0].showGridLines = True

    ws2["A1"] = "📈 Lifetime Wealth Trajectory Matrix"
    ws2["A1"].font = title_font

    headers2 = [
        "Year", "Age", "Portfolio Value", "SIP Inflows", 
        "Loan Outflows", "Goal Outflows", "Retirement Living Expenses", 
        "Total Outflows", "Phase", "Solvency"
    ]
    ws2.append([])
    ws2.append(headers2)
    h_row2 = ws2.max_row
    for col_num, h in enumerate(headers2, 1):
        cell = ws2.cell(row=h_row2, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in yearly_data:
        phase = "Retirement" if row.get("is_retirement") else "Accumulation"
        solvency = "Solvent" if row.get("is_solvent") else "Exhausted"
        ws2.append([
            row.get("year"),
            row.get("age"),
            row.get("portfolio_value", 0),
            row.get("cumulative_sip_inflows", 0),
            row.get("loan_outflow", 0),
            row.get("goal_outflow", 0),
            row.get("retirement_outflow", 0),
            row.get("annual_outflow", 0),
            phase,
            solvency
        ])
        r = ws2.max_row
        for c in range(1, 11):
            ws2.cell(row=r, column=c).font = regular_font
            ws2.cell(row=r, column=c).border = thin_border
        # Currency formatting for numeric columns
        for c in [3, 4, 5, 6, 7, 8]:
            ws2.cell(row=r, column=c).number_format = currency_fmt

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 3: Assets Inventory
    # ─────────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Assets Inventory")
    ws3.views.sheetView[0].showGridLines = True
    ws3["A1"] = "💼 Assets & Capital Holdings"
    ws3["A1"].font = title_font

    headers3 = ["Asset Name", "Category / Asset Class", "Current Value", "Return Rate (%/yr)"]
    ws3.append([])
    ws3.append(headers3)
    h_row3 = ws3.max_row
    for col_num, h in enumerate(headers3, 1):
        cell = ws3.cell(row=h_row3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for a in assets:
        ws3.append([
            a.get("name", "Asset"),
            a.get("asset_class", "other").replace("_", " ").title(),
            float(a.get("value") or 0),
            float(a.get("return_rate") or 0)
        ])
        r = ws3.max_row
        for c in range(1, 5):
            ws3.cell(row=r, column=c).font = regular_font
            ws3.cell(row=r, column=c).border = thin_border
        ws3.cell(row=r, column=3).number_format = currency_fmt
        ws3.cell(row=r, column=4).number_format = '0.0"%"'

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 4: SIPs & Investments
    # ─────────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet(title="SIPs & Investments")
    ws4.views.sheetView[0].showGridLines = True
    ws4["A1"] = "📈 Systematic Investment Plans (SIPs)"
    ws4["A1"].font = title_font

    headers4 = ["Investment Name", "Asset Class", "Monthly Amount", "Annual Contribution", "Step-Up (%/yr)", "Return Rate (%)", "Start Year", "End Year"]
    ws4.append([])
    ws4.append(headers4)
    h_row4 = ws4.max_row
    for col_num, h in enumerate(headers4, 1):
        cell = ws4.cell(row=h_row4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for s in sips:
        m_amt = float(s.get("monthly_amount") or 0)
        ws4.append([
            s.get("name", "SIP"),
            s.get("asset_class", "equity").replace("_", " ").title(),
            m_amt,
            m_amt * 12,
            float(s.get("step_up_pct") or 0),
            float(s.get("return_rate") or 0),
            int(s.get("start_year") or 2026),
            s.get("end_year") or "Indefinite"
        ])
        r = ws4.max_row
        for c in range(1, 9):
            ws4.cell(row=r, column=c).font = regular_font
            ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=3).number_format = currency_fmt
        ws4.cell(row=r, column=4).number_format = currency_fmt
        ws4.cell(row=r, column=5).number_format = '0.0"%"'
        ws4.cell(row=r, column=6).number_format = '0.0"%"'

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 5: Milestone Goals
    # ─────────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet(title="Milestone Goals")
    ws5.views.sheetView[0].showGridLines = True
    ws5["A1"] = "🎯 Milestone Financial Goals"
    ws5["A1"].font = title_font

    headers5 = ["Goal Name", "Priority", "Present Value", "Target Year", "Inflation Rate (%)", "Goal Type"]
    ws5.append([])
    ws5.append(headers5)
    h_row5 = ws5.max_row
    for col_num, h in enumerate(headers5, 1):
        cell = ws5.cell(row=h_row5, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for g in goals:
        ws5.append([
            g.get("name", "Goal"),
            g.get("priority", "need").upper(),
            float(g.get("present_value") or 0),
            int(g.get("target_year") or 2030),
            float(g.get("inflation_rate") or 6.0),
            g.get("goal_type", "lump_sum").replace("_", " ").title()
        ])
        r = ws5.max_row
        for c in range(1, 7):
            ws5.cell(row=r, column=c).font = regular_font
            ws5.cell(row=r, column=c).border = thin_border
        ws5.cell(row=r, column=3).number_format = currency_fmt
        ws5.cell(row=r, column=5).number_format = '0.0"%"'

    # ─────────────────────────────────────────────────────────────────────────────
    # SHEET 6: Loans & Insurance
    # ─────────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet(title="Loans & Insurance")
    ws6.views.sheetView[0].showGridLines = True
    ws6["A1"] = "🏦 Liabilities (Loans) & 🛡️ Insurance Policies"
    ws6["A1"].font = title_font

    ws6.append([])
    ws6.append(["🏦 Active Liabilities & Debt"])
    ws6.cell(row=ws6.max_row, column=1).font = section_font

    headers6_l = ["Loan Name", "Type", "Principal", "ROI (%/yr)", "Total Months", "EMIs Paid"]
    ws6.append(headers6_l)
    h_row6_l = ws6.max_row
    for col_num, h in enumerate(headers6_l, 1):
        cell = ws6.cell(row=h_row6_l, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for l in loans:
        ws6.append([
            l.get("name", "Loan"),
            l.get("loan_type", "home").title(),
            float(l.get("principal") or 0),
            float(l.get("roi_pct") or 0),
            int(l.get("total_months") or 240),
            int(l.get("emis_paid") or 0)
        ])
        r = ws6.max_row
        for c in range(1, 7):
            ws6.cell(row=r, column=c).font = regular_font
            ws6.cell(row=r, column=c).border = thin_border
        ws6.cell(row=r, column=3).number_format = currency_fmt
        ws6.cell(row=r, column=4).number_format = '0.0"%"'

    ws6.append([])
    ws6.append(["🛡️ Insurance Policies & Cashflows"])
    ws6.cell(row=ws6.max_row, column=1).font = section_font

    headers6_i = ["Policy Name", "Annual Premium", "Premium End Year", "Annual Income", "Income Period", "Death Benefit"]
    ws6.append(headers6_i)
    h_row6_i = ws6.max_row
    for col_num, h in enumerate(headers6_i, 1):
        cell = ws6.cell(row=h_row6_i, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ins in insurance_plans:
        inc_start = ins.get("income_start_year") or "-"
        inc_end   = ins.get("income_end_year") or "-"
        ws6.append([
            ins.get("name", "Insurance Policy"),
            float(ins.get("annual_premium") or 0),
            int(ins.get("premium_end_year") or 2035),
            float(ins.get("annual_income") or 0),
            f"{inc_start} - {inc_end}",
            float(ins.get("death_benefit") or 0)
        ])
        r = ws6.max_row
        for c in range(1, 7):
            ws6.cell(row=r, column=c).font = regular_font
            ws6.cell(row=r, column=c).border = thin_border
        ws6.cell(row=r, column=2).number_format = currency_fmt
        ws6.cell(row=r, column=4).number_format = currency_fmt
        ws6.cell(row=r, column=6).number_format = currency_fmt

    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_family_financial_excel_report(family_summary: dict) -> io.BytesIO:
    """
    Generates a consolidated Multi-Member Household Excel Workbook:
    1. Household Executive Summary (Member cards, Family KPIs, Asset Allocation)
    2. Household Trajectory (Yearly schedule per member + family total)
    3. Household Assets (With Owner / Member name column)
    4. Household SIPs (With Owner / Member name column)
    5. Household Goals (With Owner / Member name column)
    6. Household Loans & Insurance (With Owner / Member name column)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    title_font  = Font(name=font_family, size=16, bold=True, color="2D1B69")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="6B5B95")
    section_font = Font(name=font_family, size=12, bold=True, color="2D1B69")
    bold_font   = Font(name=font_family, size=11, bold=True, color="2D1B69")
    regular_font = Font(name=font_family, size=10, color="2D1B69")
    
    thin_border = Border(
        left=Side(style='thin', color='E0D7FF'),
        right=Side(style='thin', color='E0D7FF'),
        top=Side(style='thin', color='E0D7FF'),
        bottom=Side(style='thin', color='E0D7FF')
    )

    currency_fmt = '#,##0'
    pct_fmt = '0.0%'

    kpis = family_summary.get("kpis", {})
    members = family_summary.get("member_cards", [])
    allocation = family_summary.get("allocation", {})
    trajectory = family_summary.get("yearly_trajectory", [])
    assets = family_summary.get("combined_assets", [])
    goals = family_summary.get("combined_goals", [])
    sips = family_summary.get("combined_sips", [])
    insurance = family_summary.get("combined_insurance", [])
    loans = family_summary.get("combined_loans", [])

    # SHEET 1: Household Executive Summary
    ws1 = wb.create_sheet(title="Household Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    ws1["A1"] = "🔮 WealthLens — Household Master Financial Report"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Consolidated Household Analysis for {len(members)} Family Member Profiles"
    ws1["A2"].font = subtitle_font

    ws1["A4"] = "🏠 Household Key Indicators (KPIs)"
    ws1["A4"].font = section_font

    kpi_headers = ["Metric", "Value"]
    ws1.append([])
    ws1.append(kpi_headers)
    h_row = ws1.max_row
    for c, h in enumerate(kpi_headers, 1):
        cell = ws1.cell(row=h_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    kpi_rows = [
        ("Total Household Net Worth", kpis.get("net_worth", 0)),
        ("Total Household Assets Value", kpis.get("total_assets", 0)),
        ("Total Household Monthly SIPs", kpis.get("monthly_sip", 0)),
        ("Total Household Debt", kpis.get("total_debt", 0)),
        ("Household Financial Health Score", f"{kpis.get('health_score', 75)} / 100"),
        ("Linked Family Members", kpis.get("total_members", len(members))),
        ("Total Assets Count", kpis.get("total_assets_count", len(assets))),
        ("Total Active SIPs Count", kpis.get("total_sips_count", len(sips))),
    ]
    for param, val in kpi_rows:
        ws1.append([param, val])
        r = ws1.max_row
        ws1.cell(row=r, column=1).font = regular_font
        ws1.cell(row=r, column=2).font = bold_font
        if isinstance(val, (int, float)):
            ws1.cell(row=r, column=2).number_format = currency_fmt

    ws1.append([])
    ws1.append(["👤 Family Member Breakdown"])
    ws1.cell(row=ws1.max_row, column=1).font = section_font

    m_headers = ["Family Member", "Relationship / Role", "Age", "Retirement Target", "Portfolio Value (₹)", "Monthly SIP (₹)", "Assets Count", "Health Score"]
    ws1.append(m_headers)
    h_m_row = ws1.max_row
    for c, h in enumerate(m_headers, 1):
        cell = ws1.cell(row=h_m_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m in members:
        ws1.append([
            m.get("name", "Member"),
            m.get("role", "Family Member"),
            m.get("current_age", 35),
            m.get("retirement_age", 60),
            float(m.get("portfolio_value", 0)),
            float(m.get("monthly_sip", 0)),
            int(m.get("asset_count", 0)),
            f"{m.get('health_score', 70)} / 100"
        ])
        r = ws1.max_row
        for c in range(1, 9):
            ws1.cell(row=r, column=c).font = regular_font
            ws1.cell(row=r, column=c).border = thin_border
        ws1.cell(row=r, column=5).number_format = currency_fmt
        ws1.cell(row=r, column=6).number_format = currency_fmt

    # SHEET 2: Household Trajectory Schedule
    ws2 = wb.create_sheet(title="Family Trajectory Schedule")
    ws2.views.sheetView[0].showGridLines = True
    ws2["A1"] = "📊 Household Wealth Projection Timeline"
    ws2["A1"].font = title_font

    member_names = [m.get("name", "Member") for m in members]
    traj_headers = ["Year"] + [f"{name} ({m.get('role')})" for name, m in zip(member_names, members)] + ["Household Total Net Worth"]
    ws2.append([])
    ws2.append(traj_headers)
    t_row = ws2.max_row
    for c, h in enumerate(traj_headers, 1):
        cell = ws2.cell(row=t_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for tr in trajectory:
        row_vals = [tr.get("year")]
        for mname in member_names:
            row_vals.append(float(tr.get(mname, 0)))
        row_vals.append(float(tr.get("family_total", 0)))
        ws2.append(row_vals)
        r = ws2.max_row
        for c in range(1, len(row_vals) + 1):
            ws2.cell(row=r, column=c).font = regular_font
            ws2.cell(row=r, column=c).border = thin_border
            if c > 1:
                ws2.cell(row=r, column=c).number_format = currency_fmt

    # SHEET 3: Household Assets
    ws3 = wb.create_sheet(title="Household Assets")
    ws3.views.sheetView[0].showGridLines = True
    ws3["A1"] = "💼 Consolidated Household Assets Inventory"
    ws3["A1"].font = title_font

    headers3 = ["Family Member", "Role", "Asset Name", "Asset Class", "Current Value (₹)", "Expected Return Rate (%)"]
    ws3.append([])
    ws3.append(headers3)
    h_row3 = ws3.max_row
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=h_row3, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for a in assets:
        ws3.append([
            a.get("owner_name", "-"),
            a.get("owner_role", "-"),
            a.get("name", "Asset"),
            a.get("asset_class", "equity").title(),
            float(a.get("value") or 0),
            float(a.get("return_rate") or 0) / 100
        ])
        r = ws3.max_row
        for c in range(1, 7):
            ws3.cell(row=r, column=c).font = regular_font
            ws3.cell(row=r, column=c).border = thin_border
        ws3.cell(row=r, column=5).number_format = currency_fmt
        ws3.cell(row=r, column=6).number_format = pct_fmt

    # SHEET 4: Household SIPs
    ws4 = wb.create_sheet(title="Household SIPs")
    ws4.views.sheetView[0].showGridLines = True
    ws4["A1"] = "📈 Consolidated Household SIP Cashflows"
    ws4["A1"].font = title_font

    headers4 = ["Family Member", "Role", "SIP / Investment Name", "Asset Class", "Monthly Amount (₹)", "Annual Amount (₹)", "Step-Up Rate (%)", "Return Rate (%)", "Start Year", "End Year"]
    ws4.append([])
    ws4.append(headers4)
    h_row4 = ws4.max_row
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(row=h_row4, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for s in sips:
        mo = float(s.get("monthly_amount") or 0)
        end_y = s.get("end_year") or "Forever"
        ws4.append([
            s.get("owner_name", "-"),
            s.get("owner_role", "-"),
            s.get("name", "SIP"),
            s.get("asset_class", "equity").title(),
            mo,
            mo * 12,
            float(s.get("step_up_pct") or 0) / 100,
            float(s.get("return_rate") or 0) / 100,
            int(s.get("start_year") or 2026),
            end_y
        ])
        r = ws4.max_row
        for c in range(1, 11):
            ws4.cell(row=r, column=c).font = regular_font
            ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=5).number_format = currency_fmt
        ws4.cell(row=r, column=6).number_format = currency_fmt
        ws4.cell(row=r, column=7).number_format = pct_fmt
        ws4.cell(row=r, column=8).number_format = pct_fmt

    # SHEET 5: Household Goals
    ws5 = wb.create_sheet(title="Household Goals")
    ws5.views.sheetView[0].showGridLines = True
    ws5["A1"] = "🎯 Consolidated Family Milestone Goals"
    ws5["A1"].font = title_font

    headers5 = ["Family Member", "Role", "Goal Name", "Priority", "Present Value (₹)", "Target Year", "Goal Type", "Inflation Rate (%)"]
    ws5.append([])
    ws5.append(headers5)
    h_row5 = ws5.max_row
    for c, h in enumerate(headers5, 1):
        cell = ws5.cell(row=h_row5, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for g in goals:
        ws5.append([
            g.get("owner_name", "-"),
            g.get("owner_role", "-"),
            g.get("name", "Goal"),
            g.get("priority", "need").title(),
            float(g.get("present_value") or 0),
            int(g.get("target_year") or 2030),
            g.get("goal_type", "lump_sum").title(),
            float(g.get("inflation_rate") or 6) / 100
        ])
        r = ws5.max_row
        for c in range(1, 9):
            ws5.cell(row=r, column=c).font = regular_font
            ws5.cell(row=r, column=c).border = thin_border
        ws5.cell(row=r, column=5).number_format = currency_fmt
        ws5.cell(row=r, column=8).number_format = pct_fmt

    # SHEET 6: Household Insurance & Loans
    ws6 = wb.create_sheet(title="Insurance & Loans")
    ws6.views.sheetView[0].showGridLines = True
    ws6["A1"] = "🏦 Household Liabilities & Insurance Coverages"
    ws6["A1"].font = title_font

    ws6["A3"] = "🏦 Household Active Loans"
    ws6["A3"].font = section_font
    headers6_l = ["Family Member", "Role", "Loan Name", "Loan Type", "Principal (₹)", "ROI (%)", "Total Months", "EMIs Paid"]
    ws6.append([])
    ws6.append(headers6_l)
    h_row6_l = ws6.max_row
    for c, h in enumerate(headers6_l, 1):
        cell = ws6.cell(row=h_row6_l, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for l in loans:
        ws6.append([
            l.get("owner_name", "-"),
            l.get("owner_role", "-"),
            l.get("name", "Loan"),
            l.get("loan_type", "home").title(),
            float(l.get("principal") or 0),
            float(l.get("roi_pct") or 0) / 100,
            int(l.get("total_months") or 240),
            int(l.get("emis_paid") or 0)
        ])
        r = ws6.max_row
        for c in range(1, 9):
            ws6.cell(row=r, column=c).font = regular_font
            ws6.cell(row=r, column=c).border = thin_border
        ws6.cell(row=r, column=5).number_format = currency_fmt
        ws6.cell(row=r, column=6).number_format = pct_fmt

    ws6.append([])
    ws6.append(["🛡️ Household Insurance Policies"])
    ws6.cell(row=ws6.max_row, column=1).font = section_font

    headers6_i = ["Family Member", "Role", "Policy Name", "Annual Premium (₹)", "Premium End Year", "Annual Income (₹)", "Income Period", "Death Benefit (₹)"]
    ws6.append(headers6_i)
    h_row6_i = ws6.max_row
    for c, h in enumerate(headers6_i, 1):
        cell = ws6.cell(row=h_row6_i, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ins in insurance:
        inc_start = ins.get("income_start_year") or "-"
        inc_end   = ins.get("income_end_year") or "-"
        ws6.append([
            ins.get("owner_name", "-"),
            ins.get("owner_role", "-"),
            ins.get("name", "Insurance Policy"),
            float(ins.get("annual_premium") or 0),
            int(ins.get("premium_end_year") or 2035),
            float(ins.get("annual_income") or 0),
            f"{inc_start} - {inc_end}",
            float(ins.get("death_benefit") or 0)
        ])
        r = ws6.max_row
        for c in range(1, 9):
            ws6.cell(row=r, column=c).font = regular_font
            ws6.cell(row=r, column=c).border = thin_border
        ws6.cell(row=r, column=4).number_format = currency_fmt
        ws6.cell(row=r, column=6).number_format = currency_fmt
        ws6.cell(row=r, column=8).number_format = currency_fmt

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
